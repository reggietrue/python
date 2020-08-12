import openpyxl
from openpyxl import Workbook

workactual=openpyxl.load_workbook("/Volumes/EVERYBODY/ReggiePythonFolder/pythonDataDriven.xlsx")
#workactual=openpyxl.load_workbook("/Users/rtruesdale/exceldatadriven/pythonDataDriven.xlsx")
i=workactual.get_sheet_names
print(i)
sheet=workactual.active
cell=sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=4,column=2).value = "Paul"
print(sheet.cell(row=4,column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A4'].value)
print("start after this")
for i in range(1,sheet.max_row+1):
    print("-->",+sheet.max_row)
    if sheet.cell(row=i,column=1).value=="Testcase4":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)
