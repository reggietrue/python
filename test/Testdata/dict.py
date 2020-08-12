import openpyxl
from openpyxl import Workbook
Dict={}

workactual=openpyxl.load_workbook("/Volumes/EVERYBODY/ReggiePythonFolder/Dict.xlsx")
#workactual=openpyxl.load_workbook("/Users/rtruesdale/exceldatadriven/pythonDataDriven.xlsx")
i=workactual.get_sheet_names
i=workactual.get_sheet_names
sheet=workactual.active
cell=sheet.cell

for i in range(1,sheet.max_row+1):
        if sheet.cell(row=i,column=1).value=="Testcase2":
            for j in range(2,sheet.max_column+1):
                print(sheet.cell(row=1,column=j).value)
                print(sheet.cell(row=i,column=j).value)

                Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value

print(Dict)


