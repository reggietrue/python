import openpyxl
import self


class Testdata:
        test_HomePage_data = [{"Name": "Reggie Truesdale" ,"Email" : "regtruesdale@gmail.com","Passwd" :"r12eggie"},{"Name":"Patick","Email":"patrick@gmail.com","Passwd":"dlrofe"}]


        @staticmethod
        def getTestData(test_case_name1,test_case_name2):
                # = self.getLogger()
                workactual = openpyxl.load_workbook("/Volumes/EVERYBODY/ReggiePythonFolder/Dict.xlsx")
                k = workactual.get_sheet_names
                sheet = workactual.active
                cell = sheet.cell
                #for correct_test_name in range(test_case_name1,test_case_name2):
                lst = []
                for i in range(1, sheet.max_row + 1):
                        Dict = {}
                                #if sheet.cell(row=i, column=1).value == test_case_name1:
                                 # print(sheet.cell(row=i, column=1).value)
                                  #print (test_case_name1)
                                 #print(test_case_name2)
                        if sheet.cell(row=i, column=1).value == test_case_name1 or sheet.cell(row=i, column=1).value == test_case_name2:
                                print(sheet.cell(row=i, column=1).value)
                                for j in range(2,sheet.max_column + 1):
                                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                                lst.append(Dict)


                #log.info(print(lst))
                print(lst)
                return lst